"""Service xuất Excel để tạo báo cáo"""

import logging
import pandas as pd
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from app.core.config import settings

logger = logging.getLogger(__name__)

class ExcelExportService:
    """Service để tạo báo cáo Excel được định dạng"""
    
    def __init__(self):
        """Khởi tạo dịch vụ xuất Excel"""
        # Tạo thư mục tạm nếu không tồn tại
        self.temp_dir = Path(settings.EXPORT_TEMP_DIR)
        self.temp_dir.mkdir(parents=True, exist_ok=True)
    
    def export_ar_aging(self, df: pd.DataFrame, org_id: int) -> str:
        """
        Xuất báo cáo lão hóa AR sang Excel
        
        Tham số:
            df: DataFrame với dữ liệu lão hóa AR
            org_id: ID Tổ chức
            
        Trả lại:
            Đường dẫn tệp cục bộ
        """
        try:
            # Tạo tên tệp
            timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
            filename = f"ar_aging_{org_id}_{timestamp}.xlsx"
            file_path = self.temp_dir / filename
            
            # Tạo sổ làm việc Excel
            wb = Workbook()
            ws = wb.active
            ws.title = "AR Aging"
            
            # Thêm tiêu đề
            ws["A1"] = "BÁO CÁO NỢ KHÁCH (AR AGING)"
            ws["A1"].font = Font(size=14, bold=True)
            ws.merge_cells("A1:F1")
            
            # Thêm ngày
            ws["A2"] = f"Ngày: {datetime.utcnow().strftime('%d/%m/%Y')}"
            ws["A2"].font = Font(size=10)
            ws.merge_cells("A2:F2")
            
            # Thêm dữ liệu với định dạng
            self._add_dataframe_to_sheet(ws, df, start_row=4)
            
            # Điều chỉnh độ rộng cột
            ws.column_dimensions["A"].width = 15
            ws.column_dimensions["B"].width = 15
            ws.column_dimensions["C"].width = 15
            ws.column_dimensions["D"].width = 12
            ws.column_dimensions["E"].width = 12
            
            # Lưu sổ làm việc
            wb.save(file_path)
            logger.info(f"✅ Generated AR Aging report: {file_path}")
            
            return str(file_path)
            
        except Exception as e:
            logger.error(f"❌ Failed to export AR aging: {e}")
            raise
    
    def export_ap_aging(self, df: pd.DataFrame, org_id: int) -> str:
        """
        Xuất báo cáo lão hóa AP sang Excel
        
        Tham số:
            df: DataFrame với dữ liệu lão hóa AP
            org_id: ID Tổ chức
            
        Trả lại:
            Đường dẫn tệp cục bộ
        """
        try:
            # Tạo tên tệp
            timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
            filename = f"ap_aging_{org_id}_{timestamp}.xlsx"
            file_path = self.temp_dir / filename
            
            # Tạo sổ làm việc Excel
            wb = Workbook()
            ws = wb.active
            ws.title = "AP Aging"
            
            # Thêm tiêu đề
            ws["A1"] = "BÁO CÁO NỢ NHÀ CUNG CẤP (AP AGING)"
            ws["A1"].font = Font(size=14, bold=True)
            ws.merge_cells("A1:F1")
            
            # Thêm ngày
            ws["A2"] = f"Ngày: {datetime.utcnow().strftime('%d/%m/%Y')}"
            ws["A2"].font = Font(size=10)
            ws.merge_cells("A2:F2")
            
            # Thêm dữ liệu với định dạng
            self._add_dataframe_to_sheet(ws, df, start_row=4)
            
            # Điều chỉnh độ rộng cột
            ws.column_dimensions["A"].width = 15
            ws.column_dimensions["B"].width = 15
            ws.column_dimensions["C"].width = 15
            ws.column_dimensions["D"].width = 12
            ws.column_dimensions["E"].width = 12
            
            # Lưu sổ làm việc
            wb.save(file_path)
            logger.info(f"✅ Generated AP Aging report: {file_path}")
            
            return str(file_path)
            
        except Exception as e:
            logger.error(f"❌ Failed to export AP aging: {e}")
            raise
    
    def export_cashflow_forecast(self, df: pd.DataFrame, org_id: int) -> str:
        """
        Xuất dự báo dòng tiền sang Excel
        
        Tham số:
            df: DataFrame với dữ liệu dự báo
            org_id: ID Tổ chức
            
        Trả lại:
            Đường dẫn tệp cục bộ
        """
        try:
            # Tạo tên tệp
            timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
            filename = f"cashflow_forecast_{org_id}_{timestamp}.xlsx"
            file_path = self.temp_dir / filename
            
            # Tạo sổ làm việc Excel
            wb = Workbook()
            ws = wb.active
            ws.title = "Cashflow Forecast"
            
            # Thêm tiêu đề
            ws["A1"] = "DỰ BÁO DÒNG TIỀN (CASHFLOW FORECAST)"
            ws["A1"].font = Font(size=14, bold=True)
            ws.merge_cells("A1:D1")
            
            # Thêm ngày
            ws["A2"] = f"Ngày: {datetime.utcnow().strftime('%d/%m/%Y')}"
            ws["A2"].font = Font(size=10)
            ws.merge_cells("A2:D2")
            
            # Thêm dữ liệu với định dạng
            self._add_dataframe_to_sheet(ws, df, start_row=4)
            
            # Điều chỉnh độ rộng cột
            ws.column_dimensions["A"].width = 12
            ws.column_dimensions["B"].width = 18
            ws.column_dimensions["C"].width = 18
            
            # Lưu sổ làm việc
            wb.save(file_path)
            logger.info(f"✅ Generated Cashflow Forecast report: {file_path}")
            
            return str(file_path)
            
        except Exception as e:
            logger.error(f"❌ Failed to export cashflow forecast: {e}")
            raise
    
    def _add_dataframe_to_sheet(self, ws, df: pd.DataFrame, start_row: int = 1):
        """
        Thêm DataFrame vào bảng tính với định dạng
        
        Tham số:
            ws: Đối tượng bảng tính
            df: DataFrame cần thêm
            start_row: Số hàng bắt đầu
        """
        # Loại bỏ múi giờ từ các cột datetime (Excel không hỗ trợ múi giờ)
        for col in df.columns:
            if pd.api.types.is_datetime64_any_dtype(df[col]):
                df[col] = df[col].dt.tz_localize(None) if df[col].dt.tz is not None else df[col]
            elif df[col].dtype == 'object':
                # Kiểm tra các đối tượng datetime có múi giờ trong các cột object
                try:
                    df[col] = pd.to_datetime(df[col], errors='ignore')
                    if hasattr(df[col].dt, 'tz') and df[col].dt.tz is not None:
                        df[col] = df[col].dt.tz_localize(None)
                except:
                    pass
        
        # Định dạng tiêu đề
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        
        # Thêm tiêu đề
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=start_row, column=col_idx)
            cell.value = col_name
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Thêm hàng dữ liệu
        for row_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start_row + 1):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.border = border
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
                # Định dạng số với dấu phân cách hàng nghìn
                if isinstance(value, (int, float)):
                    cell.number_format = "#,##0.00"
                    cell.alignment = Alignment(horizontal="right", vertical="center")


# Thể dòng
excel_service = ExcelExportService()
